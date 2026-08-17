import datetime
import json

from flask import (
    Blueprint,
    Response,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)

from app import get_db
from app.routes.auth import admin_required

partage_bp = Blueprint("partage", __name__, url_prefix="/partage")


def _annee_par_defaut():
    return datetime.date.today().year


def _liste_annees(annee_actuelle):
    courant = datetime.date.today().year
    db = get_db()
    annees = {r["annee"] for r in db.execute("SELECT DISTINCT annee FROM partage").fetchall()}
    annees.add(courant - 2)
    annees.add(courant - 1)
    annees.add(courant)
    annees.add(courant + 1)
    annees.add(annee_actuelle)
    return sorted(a for a in annees if a)


_RENOMMER_PARTAGE = {
    "ALEM Bouira": "ALEM bouira",
    "Ain Bessem": "ALEM ain bessam",
    "Lakhdaria": "ALEM lakhdaria",
    "Sour El Ghozlane": "ALEM seg",
    "M'chedallah": "ALEM m'chedellah",
    "Bordj khris": "ALEM bordj khris",
}


def _sites_partage():
    """Sites de répartition, alignés sur l'en-tête du fichier « Consommable informatique partage.xlsx »."""
    import os

    from openpyxl import load_workbook

    from app.donnees_consommables import SITES_PARTAGE

    chemin = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "Consommable informatique partage.xlsx",
    )
    try:
        wb = load_workbook(chemin, read_only=True, data_only=True)
        feuille = wb["2024"] if "2024" in wb.sheetnames else wb[wb.sheetnames[-1]]
        for row in feuille.iter_rows(values_only=True):
            for i, v in enumerate(row):
                if v is not None and str(v).strip().lower() == "partage":
                    sites = [
                        _RENOMMER_PARTAGE.get(str(x).strip(), str(x).strip())
                        for x in row[i + 1:]
                        if x is not None and str(x).strip()
                    ]
                    if sites:
                        return sites
                    break
    except Exception:
        pass
    return list(SITES_PARTAGE)


@partage_bp.route("/")
@admin_required
def page():
    return redirect(url_for("partage.annee", annee=_annee_par_defaut()))


@partage_bp.route("/<int:annee>", methods=["GET", "POST"])
@admin_required
def annee(annee):
    db = get_db()
    if request.method == "POST":
        _enregistrer_partage(annee, request.form)
        flash("Partage " + str(annee) + " enregistré.", "success")
        return redirect(url_for("partage.annee", annee=annee))
    items = _lignes_partage(annee)
    sites = _sites_partage()
    totaux_sites = [sum(i["par_site"][j] for i in items) for j in range(len(sites))]
    stats = {
        "nb_items": len(items),
        "nb_saisis": sum(1 for i in items if i["qte_achetee"] or i["partage_total"] or i["total_sites"]),
        "total_qte": sum(i["qte_achetee"] for i in items),
        "total_partage": sum(i["partage_total"] for i in items),
        "totaux_sites": totaux_sites,
        "total_sites": sum(totaux_sites),
    }
    annees = _liste_annees(annee)
    return render_template(
        "partage/page.html",
        annee=annee,
        annees=annees,
        items=items,
        sites=sites,
        stats=stats,
    )


def _lignes_partage(annee):
    db = get_db()
    sites = _sites_partage()

    rows = db.execute(
        """SELECT pc.id AS id, pc.designation AS designation,
                  COALESCE(p.qte_achetee, 0) AS qte_achetee,
                  COALESCE(p.partage_total, 0) AS partage_total,
                  COALESCE(p.repartition, '{}') AS repartition,
                  COALESCE(p.date_maj, '') AS date_maj
           FROM partage_catalogue pc
           LEFT JOIN partage p ON p.designation_id = pc.id AND p.annee = ?
           WHERE COALESCE(pc.masque, 0) = 0
           ORDER BY pc.id""",
        (annee,),
    ).fetchall()
    items = []
    for r in rows:
        item = dict(r)
        try:
            rep = json.loads(item["repartition"]) if item["repartition"] else {}
        except (ValueError, TypeError):
            rep = {}
        item["repartition"] = rep
        item["total_sites"] = sum(int(v) for v in rep.values() if v)
        item["par_site"] = [int(rep.get(site, 0) or 0) for site in sites]
        items.append(item)
    return items


def _enregistrer_partage(annee, form):
    sites = _sites_partage()

    db = get_db()
    items = db.execute(
        "SELECT id FROM partage_catalogue WHERE COALESCE(masque, 0) = 0 ORDER BY id"
    ).fetchall()
    for it in items:
        cid = it["id"]

        def _entier(prefixe, n):
            try:
                return int(form.get(prefixe + str(n), "").strip() or 0)
            except ValueError:
                return 0

        qte = _entier("qte_", cid)
        partage_total = _entier("partage_", cid)
        rep = {}
        for i, site in enumerate(sites):
            v = _entier("site_" + str(cid) + "_", i)
            if v:
                rep[site] = v
        repartition = json.dumps(rep, ensure_ascii=False)
        db.execute(
            """INSERT INTO partage (annee, designation_id, qte_achetee, partage_total, repartition, rempli_par, date_maj)
               VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
               ON CONFLICT(annee, designation_id) DO UPDATE SET
                 qte_achetee = excluded.qte_achetee,
                 partage_total = excluded.partage_total,
                 repartition = excluded.repartition,
                 rempli_par = excluded.rempli_par,
                 date_maj = excluded.date_maj""",
            (annee, cid, qte, partage_total, repartition, session.get("nom", "")),
        )
    db.commit()


@partage_bp.route("/<int:annee>/ajouter", methods=["POST"])
@admin_required
def ajouter(annee):
    designation = request.form.get("designation", "").strip()
    if not designation:
        flash("La désignation est vide.", "danger")
        return redirect(url_for("partage.annee", annee=annee))
    db = get_db()
    existe = db.execute(
        "SELECT id FROM partage_catalogue WHERE designation = ? COLLATE NOCASE",
        (designation,),
    ).fetchone()
    if existe:
        flash("Cette désignation existe déjà.", "warning")
    else:
        db.execute("INSERT INTO partage_catalogue (designation) VALUES (?)", (designation,))
        db.commit()
        flash("Désignation ajoutée : « " + designation + " ».", "success")
    return redirect(url_for("partage.annee", annee=annee))


@partage_bp.route("/<int:annee>/importer-besoins", methods=["POST"])
@admin_required
def importer_besoins(annee):
    from app.routes.besoins import _lignes_regroupees

    groupes, sites_besoins = _lignes_regroupees(annee)
    sites_partage = _sites_partage()
    db = get_db()
    catalogue = db.execute(
        "SELECT id, designation FROM partage_catalogue WHERE COALESCE(masque, 0) = 0 ORDER BY id"
    ).fetchall()
    importes = 0
    for cat in catalogue:
        designation = cat["designation"]
        groupe = None
        for g in groupes:
            if g["groupe"].strip().lower() == designation.strip().lower():
                groupe = g
                break
        if groupe is None:
            continue
        rep = {}
        total = 0
        for site_nom, valeur in groupe["par_site"].items():
            if site_nom in sites_partage and valeur:
                rep[site_nom] = int(valeur)
                total += int(valeur)
        if not total:
            continue
        db.execute(
            """INSERT INTO partage (annee, designation_id, qte_achetee, partage_total, repartition, rempli_par, date_maj)
               VALUES (?, ?, 0, ?, ?, ?, datetime('now'))
               ON CONFLICT(annee, designation_id) DO UPDATE SET
                 partage_total = excluded.partage_total,
                 repartition = excluded.repartition,
                 rempli_par = excluded.rempli_par,
                 date_maj = excluded.date_maj""",
            (annee, cat["id"], total, json.dumps(rep, ensure_ascii=False), session.get("nom", "")),
        )
        importes += 1
    db.commit()
    flash(str(importes) + " désignation(s) importée(s) depuis les besoins " + str(annee) + ".", "success")
    return redirect(url_for("partage.annee", annee=annee))


def _lignes_export(annee):
    sites = _sites_partage()

    lignes = []
    for it in _lignes_partage(annee):
        ligne = [it["designation"], it["qte_achetee"], it["partage_total"]]
        ligne += [it["repartition"].get(site, 0) for site in sites]
        ligne.append(it["total_sites"])
        lignes.append(ligne)
    return lignes


@partage_bp.route("/export/csv/<int:annee>")
@admin_required
def exporter_csv(annee):
    import csv
    import io

    tampon = io.StringIO()
    ecrivain = csv.writer(tampon, delimiter=";")
    en_tete = ["DESIGNATION", "QTE ACHETE", "PARTAGE"]
    en_tete += _sites_partage()
    en_tete.append("TOTAL SITES")
    ecrivain.writerow(en_tete)
    for ligne in _lignes_export(annee):
        ecrivain.writerow(ligne)
    donnees = "\ufeff" + tampon.getvalue()
    return Response(
        donnees,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=partage_consommables_{}.csv".format(
                annee
            )
        },
    )


@partage_bp.route("/export/xlsx/<int:annee>")
@admin_required
def exporter_xlsx(annee):
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Partage " + str(annee)
    en_tete = ["DESIGNATION", "QTE ACHETE", "PARTAGE"]
    en_tete += _sites_partage()
    en_tete.append("TOTAL SITES")
    feuille.append(en_tete)
    for cell in feuille[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F4C81")
    for ligne in _lignes_export(annee):
        feuille.append(ligne)
    largeurs = [40, 12, 10] + [16] * len(_sites_partage()) + [12]
    for i, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(i)].width = largeur

    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)
    return send_file(
        tampon,
        as_attachment=True,
        download_name="partage_consommables_{}.xlsx".format(annee),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
