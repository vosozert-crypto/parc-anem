import csv
import datetime
import io
import subprocess
import threading

from flask import Blueprint, flash, jsonify, redirect, render_template, request, session, url_for

from app import get_db
from app.routes.auth import login_required
from app.scan import construire_plage, detecter_cidr, lister_imprimantes, ping, scan_reseau

ordinateurs_bp = Blueprint("ordinateurs", __name__, url_prefix="/ordinateurs")

CHAMPS = [
    ("nom", "Nom du poste"),
    ("numero_serie", "Numéro de série"),
    ("marque_modele", "Marque / Modèle"),
    ("processeur", "Processeur"),
    ("generation", "Génération"),
    ("ram_go", "RAM (Go)"),
    ("disque", "Disque"),
    ("arch", "Architecture OS"),
    ("user_session", "Utilisateur de session"),
    ("obs", "Observations"),
]


@ordinateurs_bp.route("/")
@login_required
def liste():
    db = get_db()
    if session.get("role") == "admin":
        machines = db.execute("SELECT * FROM machines ORDER BY nom").fetchall()
    else:
        site = session.get("site", "")
        if site:
            machines = db.execute(
                "SELECT * FROM machines WHERE site = ? ORDER BY nom", (site,)
            ).fetchall()
        else:
            machines = []
    return render_template("ordinateurs/liste.html", machines=machines)


@ordinateurs_bp.route("/ajouter", methods=["GET", "POST"])
@login_required
def ajouter():
    if request.method == "POST":
        donnees = {c[0]: request.form.get(c[0], "").strip() for c in CHAMPS}
        if not donnees["nom"]:
            flash("Le nom du poste est obligatoire.", "danger")
            return render_template("ordinateurs/formulaire.html", machine={}, titre="Ajouter un poste", champs=CHAMPS)
        db = get_db()
        db.execute(
            """INSERT INTO machines (nom, numero_serie, marque_modele, processeur,
               generation, ram_go, disque, arch, user_session, obs, site)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (donnees["nom"], donnees["numero_serie"], donnees["marque_modele"],
             donnees["processeur"], donnees["generation"], donnees["ram_go"],
             donnees["disque"], donnees["arch"], donnees["user_session"], donnees["obs"],
             session.get("site", "")),
        )
        db.commit()
        flash("Poste « " + donnees["nom"] + " » ajouté à l'inventaire.", "success")
        return redirect(url_for('ordinateurs.liste'))
    return render_template("ordinateurs/formulaire.html", machine={}, titre="Ajouter un poste", champs=CHAMPS)


@ordinateurs_bp.route("/modifier/<int:mid>", methods=["GET", "POST"])
@login_required
def modifier(mid):
    db = get_db()
    machine = db.execute("SELECT * FROM machines WHERE id = ?", (mid,)).fetchone()
    if machine is None:
        flash("Poste introuvable.", "warning")
        return redirect(url_for('ordinateurs.liste'))
    if request.method == "POST":
        donnees = {c[0]: request.form.get(c[0], "").strip() for c in CHAMPS}
        if not donnees["nom"]:
            flash("Le nom du poste est obligatoire.", "danger")
            return render_template("ordinateurs/formulaire.html", machine=machine, titre="Modifier le poste", champs=CHAMPS)
        db.execute(
            """UPDATE machines SET nom = ?, numero_serie = ?, marque_modele = ?,
               processeur = ?, generation = ?, ram_go = ?, disque = ?, arch = ?,
               user_session = ?, obs = ?
               WHERE id = ?""",
            (donnees["nom"], donnees["numero_serie"], donnees["marque_modele"],
             donnees["processeur"], donnees["generation"], donnees["ram_go"],
             donnees["disque"], donnees["arch"], donnees["user_session"], donnees["obs"], mid),
        )
        db.commit()
        flash("Poste « " + donnees["nom"] + " » modifié.", "success")
        return redirect(url_for('ordinateurs.liste'))
    return render_template("ordinateurs/formulaire.html", machine=machine, titre="Modifier le poste", champs=CHAMPS)


@ordinateurs_bp.route("/supprimer/<int:mid>", methods=["POST"])
@login_required
def supprimer(mid):
    db = get_db()
    machine = db.execute("SELECT * FROM machines WHERE id = ?", (mid,)).fetchone()
    if machine is None:
        flash("Poste introuvable.", "warning")
        return redirect(url_for('ordinateurs.liste'))
    db.execute("DELETE FROM machines WHERE id = ?", (mid,))
    db.commit()
    flash("Poste « " + machine["nom"] + " » supprimé.", "info")
    return redirect(url_for('ordinateurs.liste'))


@ordinateurs_bp.route("/scan", methods=["GET", "POST"])
@login_required
def scan():
    from app.routes.besoins import preparer_besoins_site, _site_utilisateur

    cidr_actuel = detecter_cidr() or ""
    if request.method == "POST":
        plage = request.form.get("plage", "").strip()
        try:
            parallelisme = int(request.form.get("parallelisme", "50"))
        except ValueError:
            parallelisme = 50
        if not plage:
            plage = cidr_actuel
        if not plage:
            flash("Indiquez la plage réseau (ex : 192.168.1.0/24) ; la détection automatique a échoué.", "warning")
            return render_template("ordinateurs/scan.html", cidr_actuel=cidr_actuel, parallelisme=parallelisme)
        try:
            hosts = construire_plage(plage)
        except ValueError as e:
            flash(str(e), "danger")
            return render_template("ordinateurs/scan.html", cidr_actuel=cidr_actuel, parallelisme=parallelisme)

        preparer_besoins_site(datetime.date.today().year, _site_utilisateur())
        scan_id = _demarrer_scan(plage, hosts, parallelisme)
        return redirect(url_for("ordinateurs.live", scan_id=scan_id))
    return render_template("ordinateurs/scan.html", cidr_actuel=cidr_actuel, parallelisme=50)


_SCANS = {}
_SCANS_LOCK = threading.Lock()


def _demarrer_scan(plage, hosts, parallelisme):
    import uuid

    scan_id = uuid.uuid4().hex
    etat = {
        "scan_id": scan_id,
        "plage": plage,
        "total": len(hosts),
        "scanned": 0,
        "found": 0,
        "en_cours": True,
        "erreur": None,
        "machines": [],
        "insere": False,
        "ajoutes": 0,
        "ignores": 0,
    }
    with _SCANS_LOCK:
        _SCANS[scan_id] = etat
        if len(_SCANS) > 20:
            for sid in [s for s, e in _SCANS.items() if not e["en_cours"]][: len(_SCANS) - 20]:
                _SCANS.pop(sid, None)

    def _avance(i, total, trouves, host, ok):
        etat["scanned"] = i
        etat["found"] = trouves

    def _lancer():
        try:
            machines = scan_reseau(hosts, parallelisme, progres=_avance)
            etat["machines"] = machines
        except Exception as e:
            etat["erreur"] = str(e)
        finally:
            etat["en_cours"] = False

    threading.Thread(target=_lancer, daemon=True).start()
    return scan_id


@ordinateurs_bp.route("/scan/live/<scan_id>")
@login_required
def live(scan_id):
    etat = _SCANS.get(scan_id)
    if etat is None:
        flash("Scan introuvable ou expiré.", "warning")
        return redirect(url_for('ordinateurs.scan'))
    return render_template("ordinateurs/live.html", etat=etat)


@ordinateurs_bp.route("/scan/statut/<scan_id>")
@login_required
def statut(scan_id):
    from flask import jsonify

    etat = _SCANS.get(scan_id)
    if etat is None:
        return jsonify({"erreur": "Scan introuvable."}), 404
    if not etat["en_cours"] and not etat["insere"]:
        db = get_db()
        existants = {
            r["nom"] for r in db.execute("SELECT nom FROM machines").fetchall()
        }
        for m in etat["machines"]:
            if m["nom"] in existants:
                etat["ignores"] += 1
                continue
            db.execute(
                """INSERT INTO machines (nom, numero_serie, marque_modele, processeur,
                   generation, ram_go, disque, arch, user_session, obs, site)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (m["nom"], m["numero_serie"], m["marque_modele"], m["processeur"],
                 m["generation"], m["ram_go"], m["disque"], m["arch"],
                 m.get("user_session", ""), m["obs"], session.get("site", "")),
            )
            existants.add(m["nom"])
            etat["ajoutes"] += 1
        db.commit()
        etat["insere"] = True
    return jsonify(
        {
            k: etat[k]
            for k in (
                "total", "scanned", "found", "en_cours",
                "machines", "erreur", "ajoutes", "ignores",
            )
        }
    )


def _colonnes_export():
    return [
        ("nom", "Nom du poste"),
        ("numero_serie", "Numéro de série"),
        ("marque_modele", "Marque / Modèle"),
        ("processeur", "Processeur"),
        ("generation", "Génération"),
        ("ram_go", "RAM (Go)"),
        ("disque", "Disque"),
        ("arch", "Architecture OS"),
        ("user_session", "Utilisateur de session"),
        ("obs", "Observations"),
        ("date_ajout", "Ajouté le"),
    ]


@ordinateurs_bp.route("/export/csv")
@login_required
def exporter_csv():
    from flask import Response

    db = get_db()
    machines = db.execute("SELECT * FROM machines ORDER BY nom").fetchall()
    tampon = io.StringIO()
    ecrivain = csv.writer(tampon, delimiter=";")
    ecrivain.writerow([label for _, label in _colonnes_export()])
    for m in machines:
        ecrivain.writerow([m[c] for c, _ in _colonnes_export()])
    donnees = "\ufeff" + tampon.getvalue()
    return Response(
        donnees,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=ordinateurs_anem.csv"},
    )


@ordinateurs_bp.route("/export/xlsx")
@login_required
def exporter_xlsx():
    from flask import send_file

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    db = get_db()
    machines = db.execute("SELECT * FROM machines ORDER BY nom").fetchall()

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Ordinateurs"
    entetes = [label for _, label in _colonnes_export()]
    feuille.append(entetes)
    for cell in feuille[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F4C81")
    for m in machines:
        feuille.append([m[c] for c, _ in _colonnes_export()])
    largeurs = [28, 16, 24, 36, 12, 12, 24, 14, 26, 40, 20]
    for i, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(i)].width = largeur

    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)
    return send_file(
        tampon,
        as_attachment=True,
        download_name="ordinateurs_anem.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _machine_ou_404(mid):
    db = get_db()
    machine = db.execute("SELECT * FROM machines WHERE id = ?", (mid,)).fetchone()
    if machine is None:
        return None
    return machine


@ordinateurs_bp.route("/statut/<int:mid>")
@login_required
def statut_machine(mid):
    machine = _machine_ou_404(mid)
    if machine is None:
        return jsonify({"erreur": "Poste introuvable."}), 404
    en_ligne, latence = ping(machine["nom"])
    return jsonify({"en_ligne": en_ligne, "latence_ms": latence})


@ordinateurs_bp.route("/rdp/<int:mid>", methods=["POST"])
@login_required
def rdp(mid):
    machine = _machine_ou_404(mid)
    if machine is None:
        return jsonify({"erreur": "Poste introuvable."}), 404
    try:
        subprocess.Popen(["mstsc", "/v:" + machine["nom"]])
        return jsonify({"ok": True, "message": "Bureau à distance lancé vers " + machine["nom"]})
    except Exception as e:
        return jsonify({"ok": False, "erreur": str(e)})


@ordinateurs_bp.route("/message/<int:mid>", methods=["POST"])
@login_required
def message(mid):
    machine = _machine_ou_404(mid)
    if machine is None:
        return jsonify({"erreur": "Poste introuvable."}), 404
    texte = request.form.get("texte", "").strip()
    if not texte:
        return jsonify({"ok": False, "erreur": "Message vide."}), 400
    try:
        res = subprocess.run(
            ["msg", "*", "/SERVER:" + machine["nom"], texte],
            capture_output=True,
            text=True,
            timeout=10,
        )
        if res.returncode == 0:
            return jsonify({"ok": True, "message": "Message envoyé à " + machine["nom"]})
        sortie = (res.stderr or res.stdout or "Échec d'envoi du message.").strip()
        return jsonify({"ok": False, "erreur": sortie})
    except Exception as e:
        return jsonify({"ok": False, "erreur": str(e)})


@ordinateurs_bp.route("/imprimantes/<int:mid>")
@login_required
def imprimantes(mid):
    machine = _machine_ou_404(mid)
    if machine is None:
        flash("Poste introuvable.", "warning")
        return redirect(url_for('ordinateurs.liste'))
    resultat = lister_imprimantes(machine["nom"])
    return render_template(
        "ordinateurs/imprimantes.html",
        machine=machine,
        imprimantes=resultat.get("liste", []),
        erreur=resultat.get("erreur"),
    )





