import threading

from flask import Blueprint, flash, jsonify, redirect, render_template, request, url_for

from app import get_db
from app.routes.auth import login_required
from app.scan import (
    construire_plage,
    detecter_cidr,
    scan_imprimantes_reseau,
    scanner_usb_machines,
    verifier_securite,
)

securite_bp = Blueprint("securite", __name__, url_prefix="/securite")

_CHECKS = {}
_CHECKS_LOCK = threading.Lock()


@securite_bp.route("/")
@login_required
def page():
    return render_template("securite.html")


@securite_bp.route("/imprimantes", methods=["GET", "POST"])
@login_required
def imprimantes():
    cidr_actuel = detecter_cidr() or ""
    if request.method == "POST":
        plage = request.form.get("plage", "").strip() or cidr_actuel
        if not plage:
            flash("Indiquez la plage réseau (ex : 192.168.1.0/24).", "warning")
            return render_template("securite/imprimantes.html", cidr_actuel=cidr_actuel)
        try:
            parallelisme = int(request.form.get("parallelisme", "50"))
        except ValueError:
            parallelisme = 50
        try:
            hosts = construire_plage(plage)
        except ValueError as e:
            flash(str(e), "danger")
            return render_template("securite/imprimantes.html", cidr_actuel=cidr_actuel)

        scan_id = _demarrer_scan_imprimantes(plage, hosts, parallelisme)
        return redirect(url_for("securite.live_imprimantes", scan_id=scan_id))
    return render_template("securite/imprimantes.html", cidr_actuel=cidr_actuel)


@securite_bp.route("/usb", methods=["GET", "POST"])
@login_required
def usb():
    if request.method == "POST":
        try:
            parallelisme = int(request.form.get("parallelisme", "20"))
        except ValueError:
            parallelisme = 20
        scan_id = _demarrer_scan_usb(parallelisme)
        return redirect(url_for("securite.live_usb", scan_id=scan_id))
    return render_template("securite/usb.html")





@securite_bp.route("/", methods=["POST"])
@login_required
def lancer():
    import uuid

    db = get_db()
    machines = [
        dict(m) for m in db.execute(
            "SELECT id, nom FROM machines ORDER BY nom"
        ).fetchall()
    ]
    if not machines:
        return render_template("securite.html", aucun=True)

    try:
        parallelisme = int(request.form.get("parallelisme", "50"))
    except ValueError:
        parallelisme = 50

    check_id = uuid.uuid4().hex
    etat = {
        "check_id": check_id,
        "total": len(machines),
        "checked": 0,
        "en_cours": True,
        "erreur": None,
        "machines": [],
        "resume": {"total": len(machines), "avec_kaspersky": 0, "sans": 0, "a_jour": 0, "pas_a_jour": 0, "injoignables": 0, "smb_vulnerables": 0},
    }
    with _CHECKS_LOCK:
        _CHECKS[check_id] = etat
        if len(_CHECKS) > 20:
            for cid in [c for c, e in _CHECKS.items() if not e["en_cours"]][: len(_CHECKS) - 20]:
                _CHECKS.pop(cid, None)

    def _avance(i, total, nom):
        etat["checked"] = i

    def _lancer():
        try:
            resultats = verifier_securite(machines, parallelisme, progres=_avance)
            etat["machines"] = resultats
            for m in resultats:
                r = etat["resume"]
                k = m.get("kaspersky")
                if m.get("erreur"):
                    r["injoignables"] += 1
                elif k is not None:
                    r["avec_kaspersky"] += 1
                    if k["a_jour"]:
                        r["a_jour"] += 1
                    else:
                        r["pas_a_jour"] += 1
                else:
                    r["sans"] += 1
                if m.get("smb", {}).get("vulnerable"):
                    r["smb_vulnerables"] += 1
        except Exception as e:
            etat["erreur"] = str(e)
        finally:
            etat["en_cours"] = False

    threading.Thread(target=_lancer, daemon=True).start()
    return redirect(url_for("securite.live", check_id=check_id))


@securite_bp.route("/live/<check_id>")
@login_required
def live(check_id):
    etat = _CHECKS.get(check_id)
    if etat is None:
        return redirect(url_for("securite.page"))
    return render_template("securite_live.html", etat=etat)


@securite_bp.route("/statut/<check_id>")
@login_required
def statut(check_id):
    etat = _CHECKS.get(check_id)
    if etat is None:
        return jsonify({"erreur": "Vérification introuvable."}), 404
    return jsonify(
        {
            k: etat[k]
            for k in (
                "total", "checked", "en_cours", "erreur",
                "machines", "resume",
            )
        }
    )


def _lignes_export(etat):
    """Transforme les résultats de vérification en lignes d'export."""
    lignes = []
    for m in etat["machines"]:
        k = m.get("kaspersky") or {}
        smb = m.get("smb") or {}
        erreur = m.get("erreur") or ""
        if erreur:
            install = "—"
            actif = "—"
            a_jour = "—"
        elif m.get("kaspersky") is not None:
            install = "Oui"
            actif = "Oui" if k.get("actif") else "Non"
            a_jour = "Oui" if k.get("a_jour") else "Non"
        else:
            install = "Non"
            actif = "—"
            a_jour = "—"
        if smb.get("erreur"):
            smb_etat = "—"
        elif smb.get("vulnerable"):
            smb_etat = "Vulnérable"
        elif smb.get("patche") or smb.get("smbv1") is False:
            smb_etat = "OK"
        else:
            smb_etat = "—"
        ports = ", ".join(
            (str(p.get("nom", "")) + ":" + str(p.get("port", ""))) for p in (m.get("ports") or [])
        )
        lignes.append(
            {
                "nom": m.get("nom", ""),
                "kaspersky": install,
                "actif": actif,
                "a_jour": a_jour,
                "version": k.get("version", "") or "—",
                "maj": k.get("maj", "") or "—",
                "ports": ports,
                "smb": smb_etat,
                "smbv1": "Oui" if smb.get("smbv1") else ("Non" if smb.get("smbv1") is False else "—"),
                "detail": erreur or "",
            }
        )
    return lignes


_COLONNES_EXPORT = [
    ("nom", "Poste"),
    ("kaspersky", "Kaspersky installé"),
    ("actif", "Actif"),
    ("a_jour", "À jour"),
    ("version", "Version"),
    ("maj", "Dernière MAJ"),
    ("ports", "Ports ouverts"),
    ("smb", "SMB"),
    ("smbv1", "SMBv1"),
    ("detail", "Détail / Erreur"),
]


@securite_bp.route("/export/csv/<check_id>")
@login_required
def exporter_csv(check_id):
    import csv
    import io

    from flask import Response

    etat = _CHECKS.get(check_id)
    if etat is None:
        return jsonify({"erreur": "Vérification introuvable."}), 404
    if etat["en_cours"]:
        return jsonify({"erreur": "Vérification en cours."}), 400

    tampon = io.StringIO()
    ecrivain = csv.writer(tampon, delimiter=";")
    ecrivain.writerow([label for _, label in _COLONNES_EXPORT])
    for ligne in _lignes_export(etat):
        ecrivain.writerow([ligne[c] for c, _ in _COLONNES_EXPORT])
    donnees = "\ufeff" + tampon.getvalue()
    return Response(
        donnees,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=verification_kaspersky.csv"},
    )


@securite_bp.route("/export/xlsx/<check_id>")
@login_required
def exporter_xlsx(check_id):
    import io

    from flask import send_file

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    etat = _CHECKS.get(check_id)
    if etat is None:
        return jsonify({"erreur": "Vérification introuvable."}), 404
    if etat["en_cours"]:
        return jsonify({"erreur": "Vérification en cours."}), 400

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Sécurité"
    entetes = [label for _, label in _COLONNES_EXPORT]
    feuille.append(entetes)
    for cell in feuille[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F4C81")
    for ligne in _lignes_export(etat):
        feuille.append([ligne[c] for c, _ in _COLONNES_EXPORT])
    largeurs = [24, 18, 8, 8, 14, 18, 34, 14, 8, 40]
    for i, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(i)].width = largeur

    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)
    return send_file(
        tampon,
        as_attachment=True,
        download_name="verification_kaspersky.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _demarrer_scan_imprimantes(plage, hosts, parallelisme):
    import uuid

    scan_id = uuid.uuid4().hex
    etat = {
        "type": "imprimantes",
        "scan_id": scan_id,
        "plage": plage,
        "total": len(hosts),
        "checked": 0,
        "en_cours": True,
        "erreur": None,
        "resultats": [],
    }
    with _CHECKS_LOCK:
        _CHECKS[scan_id] = etat
        if len(_CHECKS) > 30:
            for cid in [c for c, e in _CHECKS.items() if not e["en_cours"]][: len(_CHECKS) - 30]:
                _CHECKS.pop(cid, None)

    def _avance(i, total, h, trouve):
        etat["checked"] = i

    def _lancer():
        try:
            etat["resultats"] = scan_imprimantes_reseau(hosts, parallelisme, progres=_avance)
        except Exception as e:
            etat["erreur"] = str(e)
        finally:
            etat["en_cours"] = False

    threading.Thread(target=_lancer, daemon=True).start()
    return scan_id


def _demarrer_scan_usb(parallelisme):
    import uuid

    db = get_db()
    machines = [
        dict(m) for m in db.execute(
            "SELECT id, nom FROM machines ORDER BY nom"
        ).fetchall()
    ]
    scan_id = uuid.uuid4().hex
    etat = {
        "type": "usb",
        "scan_id": scan_id,
        "total": len(machines),
        "checked": 0,
        "en_cours": True,
        "erreur": None,
        "machines": [],
    }
    with _CHECKS_LOCK:
        _CHECKS[scan_id] = etat
        if len(_CHECKS) > 30:
            for cid in [c for c, e in _CHECKS.items() if not e["en_cours"]][: len(_CHECKS) - 30]:
                _CHECKS.pop(cid, None)

    def _avance(i, total, nom):
        etat["checked"] = i

    def _lancer():
        try:
            etat["machines"] = scanner_usb_machines(machines, parallelisme, progres=_avance)
        except Exception as e:
            etat["erreur"] = str(e)
        finally:
            etat["en_cours"] = False

    threading.Thread(target=_lancer, daemon=True).start()
    return scan_id


@securite_bp.route("/imprimantes/live/<scan_id>")
@login_required
def live_imprimantes(scan_id):
    etat = _CHECKS.get(scan_id)
    if etat is None:
        return redirect(url_for("securite.imprimantes"))
    return render_template("securite/imprimantes_live.html", etat=etat)


@securite_bp.route("/imprimantes/statut/<scan_id>")
@login_required
def statut_imprimantes(scan_id):
    etat = _CHECKS.get(scan_id)
    if etat is None:
        return jsonify({"erreur": "Scan introuvable."}), 404
    return jsonify(
        {
            k: etat[k]
            for k in ("total", "checked", "en_cours", "erreur", "resultats", "plage")
        }
    )


@securite_bp.route("/usb/live/<scan_id>")
@login_required
def live_usb(scan_id):
    etat = _CHECKS.get(scan_id)
    if etat is None:
        return redirect(url_for("securite.usb"))
    return render_template("securite/usb_live.html", etat=etat)


@securite_bp.route("/usb/statut/<scan_id>")
@login_required
def statut_usb(scan_id):
    etat = _CHECKS.get(scan_id)
    if etat is None:
        return jsonify({"erreur": "Scan introuvable."}), 404
    return jsonify(
        {
            k: etat[k]
            for k in ("total", "checked", "en_cours", "erreur", "machines")
        }
    )
