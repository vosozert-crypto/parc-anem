import datetime

from flask import (
    Blueprint,
    Response,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)

from app import get_db
from app.routes.auth import admin_required, login_required

besoins_bp = Blueprint("besoins", __name__, url_prefix="/besoins")

ETATS_IMPRIMANTE = ["neuf", "moyenne", "dégradé", "en panne", "autre"]


def _annee_par_defaut():
    return datetime.date.today().year


def _liste_annees(annee_actuelle):
    courant = datetime.date.today().year
    db = get_db()
    annees = {r["annee"] for r in db.execute("SELECT DISTINCT annee FROM besoins").fetchall()}
    annees.add(courant - 2)
    annees.add(courant - 1)
    annees.add(courant)
    annees.add(courant + 1)
    annees.add(annee_actuelle)
    return sorted(a for a in annees if a)


def _site_utilisateur():
    db = get_db()
    ligne = db.execute(
        "SELECT site FROM users WHERE id = %s", (session["user_id"],)
    ).fetchone()
    return (ligne["site"] or "").strip() if ligne else ""


def preparer_besoins_site(annee, site):
    """Crée automatiquement les lignes besoins du site pour l'année (sans écraser les saisies).

    À appeler lors d'un scan (PC / imprimantes) : chaque consommable du site
    reçoit une ligne besoins vide, afin que le formulaire et la consolidation
    les affichent sans que le collaborateur ait à soumettre le formulaire.
    """
    if not site:
        return 0
    db = get_db()
    items = db.execute(
        "SELECT id FROM consommables WHERE site = %s ORDER BY id", (site,)
    ).fetchall()
    nb = 0
    for it in items:
        cur = db.execute(
            """INSERT INTO besoins
                 (consommable_id, annee, stock, besoin, etat, rempli_par, date_maj)
               VALUES (%s, %s, '', 0, '', 'scan automatique', NOW())
               ON CONFLICT(consommable_id, annee) DO NOTHING""",
            (it["id"], annee),
        )
        nb += cur.rowcount
    db.commit()
    return nb


def _autorise_site(site):
    if session.get("role") == "admin":
        return True
    return _site_utilisateur() == site


@besoins_bp.route("/")
@login_required
def page():
    annee = _annee_par_defaut()
    req = request.args.get("annee")
    if req:
        try:
            annee = int(req)
        except ValueError:
            pass
    if session.get("role") == "admin":
        return redirect(url_for("besoins.annee", annee=annee))
    site = _site_utilisateur()
    if not site:
        return render_template("besoins/page.html", annee=annee, role="user", site=None, lignes=[])
    return redirect(url_for("besoins.formulaire", annee=annee, site=site))


@besoins_bp.route("/<int:annee>", methods=["GET", "POST"])
@login_required
def annee(annee):
    if session.get("role") != "admin":
        site = _site_utilisateur()
        if site:
            return redirect(url_for("besoins.formulaire", annee=annee, site=site))
        return redirect(url_for("besoins.page"))
    db = get_db()
    if request.method == "POST":
        site_filtre = request.form.get("site", "").strip()
        _enregistrer_besoins(annee, request.form, site=site_filtre or None)
        flash("Besoins " + str(annee) + " enregistrés.", "success")
        return redirect(
            url_for("besoins.annee", annee=annee, site=site_filtre or None)
        )
    site_filtre = request.args.get("site", "").strip()
    items = _lignes_consolidees(annee, site_filtre or None)
    groupes, sites_groupes = _lignes_regroupees(annee, site_filtre or None)
    totaux_groupes = [
        sum(g["par_site"].get(s, 0) for g in groupes) for s in sites_groupes
    ]
    sites = {i["site"] for i in items}
    saisis = {i["site"] for i in items if i["stock"] or i["besoin"] or i["etat"]}
    stats = {"nb_sites": len(sites), "nb_items": len(items), "nb_saisis": len(saisis)}
    annees = _liste_annees(annee)
    liste_sites = [
        r["site"] for r in db.execute("SELECT DISTINCT site FROM consommables ORDER BY site")
    ]
    return render_template(
        "besoins/page.html",
        annee=annee,
        annees=annees,
        role="admin",
        items=items,
        stats=stats,
        groupes=groupes,
        sites_groupes=sites_groupes,
        totaux_groupes=totaux_groupes,
        liste_sites=liste_sites,
        site_filtre=site_filtre,
        etats=ETATS_IMPRIMANTE,
        site=None,
    )


def _lignes_consolidees(annee, site=None):
    db = get_db()
    requete = """SELECT c.id AS id, c.site AS site, c.designation AS designation,
                  c.ref_toner AS ref_toner,
                  COALESCE(b.stock, '') AS stock,
                  COALESCE(b.besoin, 0) AS besoin,
                  COALESCE(b.etat, '') AS etat,
                  TO_CHAR(b.date_maj, 'YYYY-MM-DD HH24:MI:SS') AS date_maj
           FROM consommables c
           LEFT JOIN besoins b ON b.consommable_id = c.id AND b.annee = %s"""
    params = [annee]
    if site:
        requete += " WHERE c.site = %s"
        params.append(site)
    requete += " ORDER BY c.site, c.id"
    rows = db.execute(requete, params).fetchall()
    lignes = []
    numero = 0
    dernier_site = None
    for r in rows:
        if r["site"] != dernier_site:
            numero = 0
            dernier_site = r["site"]
        numero += 1
        ligne = dict(r)
        ligne["numero"] = numero
        lignes.append(ligne)
    return lignes


def _lignes_regroupees(annee, site=None):
    """Besoins regroupés par référence toner (type d'imprimante), avec colonnes par site."""
    db = get_db()
    requete = """SELECT COALESCE(NULLIF(c.ref_toner, ''), 'Non renseigné') AS groupe,
                  c.site AS site,
                  c.designation AS designation,
                  SUM(COALESCE(b.besoin, 0)) AS total
           FROM consommables c
           LEFT JOIN besoins b ON b.consommable_id = c.id AND b.annee = %s"""
    params = [annee]
    if site:
        requete += " WHERE c.site = %s"
        params.append(site)
    requete += " GROUP BY groupe, c.site, c.designation ORDER BY groupe, c.site, c.designation"
    rows = db.execute(requete, params).fetchall()
    groupes = {}
    sites = {}
    for r in rows:
        g = groupes.setdefault(
            r["groupe"], {"groupe": r["groupe"], "par_site": {}, "designations": set(), "total": 0}
        )
        g["par_site"][r["site"]] = g["par_site"].get(r["site"], 0) + r["total"]
        g["total"] += r["total"]
        g["designations"].add(r["designation"])
        sites[r["site"]] = True
    liste = list(groupes.values())
    for g in liste:
        g["designations"] = sorted(g["designations"])
        g["nb_sites"] = len(g["par_site"])
    liste.sort(key=lambda g: (-g["total"], g["groupe"]))
    return liste, sorted(sites.keys())


def _enregistrer_besoins(annee, form, site=None):
    db = get_db()
    if site:
        items = db.execute(
            "SELECT id FROM consommables WHERE site = %s ORDER BY id", (site,)
        ).fetchall()
    else:
        items = db.execute(
            "SELECT id FROM consommables ORDER BY site, id"
        ).fetchall()
    for it in items:
        cid = it["id"]
        stock = form.get("stock_" + str(cid), "").strip()
        etat = form.get("etat_" + str(cid), "").strip()
        try:
            besoin = int(form.get("besoin_" + str(cid), "").strip() or 0)
        except ValueError:
            besoin = 0
        db.execute(
            """INSERT INTO besoins
                 (consommable_id, annee, stock, besoin, etat, rempli_par, date_maj)
               VALUES (%s, %s, %s, %s, %s, %s, NOW())
               ON CONFLICT(consommable_id, annee) DO UPDATE SET
                 stock = excluded.stock,
                 besoin = excluded.besoin,
                 etat = excluded.etat,
                 rempli_par = excluded.rempli_par,
                 date_maj = excluded.date_maj""",
            (cid, annee, stock, besoin, etat, session.get("nom", "")),
        )
        ref = form.get("ref_" + str(cid), "").strip()
        if ref:
            db.execute(
                "UPDATE consommables SET ref_toner = %s WHERE id = %s", (ref, cid)
            )
    db.commit()


@besoins_bp.route("/statut/<int:annee>")
@admin_required
def statut(annee):
    from flask import jsonify

    db = get_db()
    site = request.args.get("site", "").strip()
    requete = """SELECT c.id AS id, c.site AS site, c.designation AS designation,
                  c.ref_toner AS ref_toner,
                  COALESCE(b.stock, '') AS stock,
                  COALESCE(b.besoin, 0) AS besoin,
                  COALESCE(b.etat, '') AS etat,
                  TO_CHAR(b.date_maj, 'YYYY-MM-DD HH24:MI:SS') AS date_maj
           FROM consommables c
           LEFT JOIN besoins b ON b.consommable_id = c.id AND b.annee = %s"""
    params = [annee]
    if site:
        requete += " WHERE c.site = %s"
        params.append(site)
    requete += " ORDER BY c.site, c.id"
    rows = db.execute(requete, params).fetchall()
    return jsonify([dict(r) for r in rows])


@besoins_bp.route("/<int:annee>/<site>", methods=["GET", "POST"])
@login_required
def formulaire(annee, site):
    if not _autorise_site(site):
        flash("Vous n'avez pas accès au site « " + site + " ».", "danger")
        return redirect(url_for("besoins.page"))
    db = get_db()
    if request.method == "POST":
        _enregistrer_besoins(annee, request.form, site=site)
        flash(
            "Besoins " + str(annee) + " du site « " + site + " » enregistrés.",
            "success",
        )
        return redirect(url_for("besoins.formulaire", annee=annee, site=site))
    items = db.execute(
        """SELECT c.id AS id, c.designation AS designation,
                  c.ref_toner AS ref_toner,
                  COALESCE(b.stock, '') AS stock,
                  COALESCE(b.besoin, 0) AS besoin,
                  COALESCE(b.etat, '') AS etat,
                  TO_CHAR(b.date_maj, 'YYYY-MM-DD HH24:MI:SS') AS date_maj
           FROM consommables c
           LEFT JOIN besoins b ON b.consommable_id = c.id AND b.annee = %s
           WHERE c.site = %s
           ORDER BY c.id""",
        (annee, site),
    ).fetchall()
    return render_template(
        "besoins/formulaire.html",
        annee=annee,
        site=site,
        items=items,
        role=session.get("role"),
        etats=ETATS_IMPRIMANTE,
    )


def _lignes_export(annee, site=None):
    db = get_db()
    requete = """SELECT c.site AS site, c.designation AS designation,
                  COALESCE(b.stock, '') AS stock,
                  COALESCE(b.besoin, 0) AS besoin,
                  COALESCE(b.etat, '') AS etat
           FROM consommables c
           LEFT JOIN besoins b ON b.consommable_id = c.id AND b.annee = %s"""
    params = [annee]
    if site:
        requete += " WHERE c.site = %s"
        params.append(site)
    requete += " ORDER BY c.site, c.id"
    rows = db.execute(requete, params).fetchall()
    lignes = []
    site_actuel = None
    numero = 0
    for r in rows:
        if r["site"] != site_actuel:
            site_actuel = r["site"]
            numero = 0
        numero += 1
        lignes.append(
            [
                r["site"] if numero == 1 else "",
                numero,
                r["designation"],
                r["stock"],
                r["besoin"],
                r["etat"],
                "",
            ]
        )
    groupes, _ = _lignes_regroupees(annee, site)
    lignes.append([""] * 7)
    lignes.append(
        ["Consolidé Total", "N°", "Référence (Tonner/Cartouche)", "", "BESOIN Total", "", "prix en da"]
    )
    numero = 0
    for g in groupes:
        numero += 1
        designation = g["designations"][0] if g["designations"] else g["groupe"]
        lignes.append(["", numero, designation, "", g["total"], "", ""])
    return lignes


EN_TETE_EXPORT = ["Site", "N°", "Référence (Tonner/Cartouche)", "STOCK", "BESOIN", "ETAT de Tonner/", ""]


@besoins_bp.route("/export/csv/<int:annee>")
@admin_required
def exporter_csv(annee):
    import csv
    import io

    site = request.args.get("site", "").strip() or None
    tampon = io.StringIO()
    ecrivain = csv.writer(tampon, delimiter=";")
    ecrivain.writerow(EN_TETE_EXPORT)
    for ligne in _lignes_export(annee, site):
        ecrivain.writerow(ligne)
    donnees = "\ufeff" + tampon.getvalue()
    return Response(
        donnees,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=besoin_consommables_{}.csv".format(
                annee
            )
        },
    )


@besoins_bp.route("/export/xlsx/<int:annee>")
@admin_required
def exporter_xlsx(annee):
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    site = request.args.get("site", "").strip() or None
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Consolidé"
    feuille.append(EN_TETE_EXPORT)
    for cell in feuille[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F4C81")
    for ligne in _lignes_export(annee, site):
        feuille.append(ligne)
    for i, largeur in enumerate([22, 6, 52, 14, 10, 16, 12], start=1):
        feuille.column_dimensions[get_column_letter(i)].width = largeur

    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)
    return send_file(
        tampon,
        as_attachment=True,
        download_name="besoin_consommables_{}.xlsx".format(annee),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _lignes_export_type(annee, site=None):
    """Besoins regroupés par réf. toner, avec une colonne par site."""
    groupes, sites_groupes = _lignes_regroupees(annee, site)
    lignes = []
    for g in groupes:
        ligne = [g["groupe"]]
        ligne += [g["par_site"].get(s, 0) for s in sites_groupes]
        ligne.append(g["total"])
        lignes.append(ligne)
    totaux = [
        sum(g["par_site"].get(s, 0) for g in groupes) for s in sites_groupes
    ]
    lignes.append(
        ["Total"] + totaux + [sum(g["total"] for g in groupes)]
    )
    return lignes, sites_groupes


@besoins_bp.route("/export/type/csv/<int:annee>")
@admin_required
def exporter_csv_type(annee):
    import csv
    import io

    site = request.args.get("site", "").strip() or None
    lignes, sites = _lignes_export_type(annee, site)
    tampon = io.StringIO()
    ecrivain = csv.writer(tampon, delimiter=";")
    ecrivain.writerow(["Réf. toner / type imprimante"] + sites + ["Total"])
    for ligne in lignes:
        ecrivain.writerow(ligne)
    donnees = "\ufeff" + tampon.getvalue()
    return Response(
        donnees,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=besoin_par_type_{}.csv".format(
                annee
            )
        },
    )


@besoins_bp.route("/export/type/xlsx/<int:annee>")
@admin_required
def exporter_xlsx_type(annee):
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    site = request.args.get("site", "").strip() or None
    lignes, sites = _lignes_export_type(annee, site)
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Par type imprimante " + str(annee)
    en_tete = ["Réf. toner / type imprimante"] + sites + ["Total"]
    feuille.append(en_tete)
    for cell in feuille[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F4C81")
    for ligne in lignes:
        feuille.append(ligne)
    largeurs = [42] + [12] * len(sites) + [10]
    for i, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(i)].width = largeur

    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)
    return send_file(
        tampon,
        as_attachment=True,
        download_name="besoin_par_type_{}.xlsx".format(annee),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

